import os
import openpyxl
import math
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart,ScatterChart, Series, Reference


os.system("cls")

example_0 = [82,85,86,87,87,89,89,90,91,91,92,93,94,95,95,95,95,95,97,98,99,99,100,100,101,101,103,103,103,104,105,105,106,107,107,107,109,110,110,111]
example_1 = [2.97,4.00,5.20,5.56,5.94,5.98,6.35,6.62,6.72,6.78,6.80,6.85,6.94,7.15,7.16,7.23,7.29,7.62,7.62,7.73,7.87,7.93,8.00,8.26,8.29,8.37,8.47,8.54,8.54,8.58,8.61,8.67,8.69,8.81,9.07,9.27,9.37,9.43,9.52,9.58,9.60,9.76,9.82,9.83,9.83,9.84,9.96,10.04,10.21,10.28,10.28,10.30,10.35,10.36,10.40,10.49,10.49,10.50,10.64,10.95,11.09,11.12,11.21,11.29,11.43,11.62,11.70,12.16,12.19,12.28,12.31,12.62,12.69,12.71,12.91,12.92,13.11,13.38,13.42,13.43,13.47,13.60,13.96,14.24,14.35,15.12,15.24,16.06,16.90,18.26]
example_2 = [11.5,12.1,9.9,9.3,7.8,6.2,6.6,7.0,13.4,17.1,9.3,5.6,5.7,5.4,5.2,5.1,4.9,10.7,15.2,8.5,4.2,4.0,3.9,3.8,3.6,3.4,20.6,25.5,13.8,12.6,13.1,8.9,8.2,10.7,14.2,7.6,5.2,5.5,5.1,5.0,5.2,4.8,4.1,3.8,3.7,3.6,3.6,3.6]
def data(array):
  total = len(array)
  array.sort()
  interval = array[0] - array[total-1]
  #print(array)

  print("total:"+str(total))
  print("interval:" +str(abs(interval)))

  classNumber = 6
  ValueNumber = 5
  off = 0
  initial = 0
  y = 0
  actual = 0
  same = 0

  classInterval = []
  middlePoint = []
  classFrequency = []
  relativeFrequency = []
  display = []
  infoGeogebra = []

  for i in range(classNumber):
    if off>0:
      initial = 0.1

    parameter = (array[0]+off+initial),(array[0]+ValueNumber+off)
    classInterval.append(parameter)
    off = off + ValueNumber

  for i in classInterval:
    middlePoint.append((i[0]+i[1])/2)
    count = 0
    for j in array:
      if( j >= i[0] and j<=i[1]):
        if(actual == j ):
          y = y + 0.4
          same = same + 0.4
        if(actual != j):
          y = y - same
          same = 0
        
        tupleGeo = (j,y)
        infoGeogebra.append(tupleGeo)
        count = count + 1
        actual = j
    y = y + 1
    classFrequency.append(count)
  
  for i in classFrequency:
    relativeFrequency.append(i/total)

  for i in range(classNumber):
    tuple = (str(classInterval[i]),relativeFrequency[i],classFrequency[i],middlePoint[i])
    display.append(tuple)

  wb = openpyxl.Workbook()
  hoja = wb.active
  hoja.append(["Class Interval","Relative Frequencys","Class Frequency","Middle Point"])
  for x in display:
    hoja.append(x)

  hoja.cell(column = 10 , row = 1 ,value= "x")
  hoja.cell(column = 11 , row = 1 ,value= "y")
  for row,value in enumerate(infoGeogebra,start = 2):
    hoja.cell(column = 10 , row = row ,value= value[0])
    hoja.cell(column = 11 , row = row ,value= value[1])

  
  ref = "A1:" + "D" + str(classNumber + 1) 
  ref_2 = "J1:" + "K" + str(hoja.max_row)

  tab = Table(displayName="Table1", ref=ref)
  tab_2 = Table(displayName="Table2", ref=ref_2)
  style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
  tab.tableStyleInfo = style
  tab_2.tableStyleInfo = style

 
  hoja.add_table(tab)
  hoja.add_table(tab_2)

  chart1 = BarChart()
  chart1.type = "col"
  chart1.style = 10
  chart1.title = "Histogram"
  chart1.y_axis.title = 'Relative Frequency'
  chart1.x_axis.title = 'Class Interval'
  data = Reference(hoja, min_col=2, min_row=1, max_row=classNumber+1, max_col=2)
  cats = Reference(hoja, min_col=1, min_row=2, max_row=classNumber+1)
  chart1.add_data(data, titles_from_data=True)
  chart1.set_categories(cats)
  chart1.shape = 4
  hoja.add_chart(chart1, "A10")


  chart = ScatterChart()
  chart.title = "Dot Plot"
  chart.style = 12
  chart.x_axis.title = ''
  chart.y_axis.title = ''
  chart.x_axis.scaling.min = array[0]
  chart.x_axis.scaling.max = array[total - 1 ]
  chart.legend = None

  xvalues = Reference(hoja, min_col = 10, min_row = 2, max_row = hoja.max_row)
  values = Reference(hoja, min_col = 11, min_row = 1, max_row = hoja.max_row)
  series = Series(values, xvalues, title_from_data = True)
  series.marker.symbol = "circle"
  series.graphicalProperties.line.noFill = True
  chart.y_axis.title = ""
  chart.series.append(series)
  hoja.add_chart(chart, "M10")


  wb.save('pyhton.xlsx')

  
if __name__ == "__main__":
  data(example_0)
